"""XLSX handler.

Only cell values are replaced. Number formats, styles, column widths, merges,
conditional formatting and formulas all live separately from the value, so they
survive untouched. Formula cells are skipped deliberately: translating ``=SUM(A1)``
would break the sheet, and the cached result is recomputed by Excel anyway.
"""

import io
from collections.abc import Iterator
from typing import Any

from src.api.services.document_format.base import (
    DocumentFormatError,
    FormatHandler,
    TextUnit,
)

# Cells that hold a string. openpyxl reports formulas as "f" and numbers,
# booleans, errors and dates with their own type codes.
_STRING_TYPE = "s"


class _CellUnit:
    def __init__(self, cell: Any, sheet_title: str) -> None:
        self._cell = cell
        self._context = f"{sheet_title}!{cell.coordinate}"

    @property
    def context(self) -> str:
        return self._context

    def get_text(self) -> str:
        value = self._cell.value
        return value if isinstance(value, str) else ""

    def set_text(self, value: str) -> None:
        self._cell.value = value


class XlsxHandler(FormatHandler):
    extensions = ("xlsx",)
    output_extension = ".xlsx"
    output_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def load(self, data: bytes) -> Any:
        from openpyxl import load_workbook

        try:
            # data_only=False keeps formulas as formulas instead of freezing them
            # into their last cached value.
            return load_workbook(io.BytesIO(data), data_only=False)
        except Exception as exc:
            raise DocumentFormatError(f"Could not read XLSX file: {exc}") from exc

    def save(self, document: Any) -> bytes:
        buffer = io.BytesIO()
        document.save(buffer)
        return buffer.getvalue()

    def units(self, document: Any) -> Iterator[TextUnit]:
        for sheet in document.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type != _STRING_TYPE:
                        continue
                    if not isinstance(cell.value, str) or cell.value.startswith("="):
                        continue
                    yield _CellUnit(cell, sheet.title)
